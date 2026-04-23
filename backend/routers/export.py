import io
from datetime import date
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.services.data_loader import (
    get_products,
    get_categories,
    get_orders,
    get_customers,
    get_reviews,
    get_sales_summary,
)

router = APIRouter(prefix="/export", tags=["export"])

# ── Style constants ───────────────────────────────────────────────────────────

_INDIGO = "4F46E5"
_INDIGO_LIGHT = "E0E7FF"
_WHITE = "FFFFFF"
_DARK = "1E1B4B"
_BORDER_COLOR = "D1D5DB"

_THIN = Side(style="thin", color=_BORDER_COLOR)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HDR_FONT = Font(bold=True, color=_WHITE, name="Arial", size=10)
_HDR_FILL = PatternFill("solid", start_color=_INDIGO)
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_BODY_FONT = Font(name="Arial", size=9)
_BODY_BOLD = Font(name="Arial", size=9, bold=True)
_BODY_ALIGN = Alignment(vertical="center")


def _hdr(cell):
    cell.font = _HDR_FONT
    cell.fill = _HDR_FILL
    cell.alignment = _HDR_ALIGN
    cell.border = _BORDER


def _body(cell, bold=False):
    cell.font = _BODY_BOLD if bold else _BODY_FONT
    cell.alignment = _BODY_ALIGN
    cell.border = _BORDER


def _auto_width(ws, mn=10, mx=45):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        best = max((len(str(c.value or "")) for c in col), default=mn)
        ws.column_dimensions[letter].width = min(max(mn, best + 2), mx)


# ── Endpoint ──────────────────────────────────────────────────────────────────

@router.get("/excel", summary="Download full store data as Excel")
def export_excel():
    """Export all store data (products, orders, customers, reviews, sales) as a
    multi-sheet Excel workbook."""

    products = get_products()
    categories = get_categories()
    orders = get_orders()
    customers = get_customers()
    reviews = get_reviews()
    sales = get_sales_summary()

    cat_map = {c["id"]: c["name"] for c in categories}
    prod_map = {p["id"]: p["name"] for p in products}
    cust_map = {c["id"]: c["name"] for c in customers}

    wb = Workbook()
    wb.remove(wb.active)  # drop default blank sheet

    # ── Sheet 1 : Overview ────────────────────────────────────────────────────
    ws = wb.create_sheet("Overview")

    ws.merge_cells("A1:C1")
    ws["A1"] = "AskMyStore — Full Data Export"
    ws["A1"].font = Font(bold=True, name="Arial", size=16, color=_INDIGO)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Generated: {date.today().strftime('%B %d, %Y')}"
    ws["A2"].font = Font(name="Arial", size=10, color="6B7280")

    ws.append([])  # blank row

    completed = [o for o in orders if o.get("status") == "completed"]
    return_reqs = [r for r in reviews if r.get("has_return_request")]

    summary_rows = [
        ("Total Products", len(products)),
        ("Product Categories", len(categories)),
        ("Total Orders", len(orders)),
        ("Completed Orders", len(completed)),
        ("Processing / Shipped Orders",
         len([o for o in orders if o.get("status") in ("processing", "shipped")])),
        ("Cancelled Orders",
         len([o for o in orders if o.get("status") == "cancelled"])),
        ("Total Customers", len(customers)),
        ("Total Reviews", len(reviews)),
        ("Total Return Requests", len(return_reqs)),
        ("Return Rate",
         f"{len(return_reqs) / len(reviews) * 100:.1f}%" if reviews else "—"),
        ("Export Date", str(date.today())),
    ]

    ws.append(["Metric", "Value"])
    hdr_row = ws.max_row
    for col in range(1, 3):
        _hdr(ws.cell(hdr_row, col))
    ws.row_dimensions[hdr_row].height = 20

    for label, value in summary_rows:
        ws.append([label, value])
        r = ws.max_row
        _body(ws.cell(r, 1), bold=True)
        _body(ws.cell(r, 2))

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    # ── Sheet 2 : Products ────────────────────────────────────────────────────
    ws = wb.create_sheet("Products")
    headers = ["ID", "Name", "Category", "Price ($)", "Cost ($)",
               "Margin (%)", "Stock", "Rating", "Reviews Count", "Description"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        _hdr(ws.cell(1, col))
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for p in products:
        price = p.get("price", 0) or 0
        cost = p.get("cost", 0) or 0
        margin = round((price - cost) / price * 100, 1) if price else 0
        row = [
            p["id"],
            p.get("name", ""),
            cat_map.get(p.get("category_id"), "Unknown"),
            round(price, 2),
            round(cost, 2),
            margin,
            p.get("stock_quantity", p.get("stock", 0)),
            p.get("rating", ""),
            p.get("reviews_count", ""),
            p.get("description", ""),
        ]
        ws.append(row)
        r = ws.max_row
        for col in range(1, len(headers) + 1):
            _body(ws.cell(r, col), bold=(col == 1))

    _auto_width(ws)

    # ── Sheet 3 : Orders ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Orders")
    headers = ["Order ID", "Date", "Customer", "Items Summary",
               "Total ($)", "Shipping ($)", "Status"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        _hdr(ws.cell(1, col))
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for o in sorted(orders, key=lambda x: x.get("order_date", ""), reverse=True):
        items_str = "; ".join(
            f"{prod_map.get(i['product_id'], 'Unknown')} ×{i['quantity']}"
            for i in o.get("items", [])
        )
        row = [
            o["id"],
            o.get("order_date", ""),
            cust_map.get(o.get("customer_id"), "Unknown"),
            items_str,
            round(o.get("total", 0), 2),
            round(o.get("shipping_cost", 0), 2),
            o.get("status", ""),
        ]
        ws.append(row)
        r = ws.max_row
        for col in range(1, len(headers) + 1):
            _body(ws.cell(r, col), bold=(col == 1))

    _auto_width(ws)

    # ── Sheet 4 : Order Items (flat) ──────────────────────────────────────────
    ws = wb.create_sheet("Order Items")
    headers = ["Order ID", "Order Date", "Customer", "Order Status",
               "Product", "Qty", "Unit Price ($)", "Line Total ($)"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        _hdr(ws.cell(1, col))
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for o in sorted(orders, key=lambda x: x.get("order_date", ""), reverse=True):
        for item in o.get("items", []):
            pid = item.get("product_id")
            qty = item.get("quantity", 0)
            unit_price = item.get("unit_price", 0)
            row = [
                o["id"],
                o.get("order_date", ""),
                cust_map.get(o.get("customer_id"), "Unknown"),
                o.get("status", ""),
                prod_map.get(pid, "Unknown"),
                qty,
                round(unit_price, 2),
                round(unit_price * qty, 2),
            ]
            ws.append(row)
            r = ws.max_row
            for col in range(1, len(headers) + 1):
                _body(ws.cell(r, col), bold=(col == 1))

    _auto_width(ws)

    # ── Sheet 5 : Customers ───────────────────────────────────────────────────
    ws = wb.create_sheet("Customers")
    headers = ["ID", "Name", "Email", "City", "State", "Join Date"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        _hdr(ws.cell(1, col))
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for c in customers:
        row = [
            c["id"],
            c.get("name", ""),
            c.get("email", ""),
            c.get("city", ""),
            c.get("state", ""),
            c.get("join_date", ""),
        ]
        ws.append(row)
        r = ws.max_row
        for col in range(1, len(headers) + 1):
            _body(ws.cell(r, col), bold=(col == 1))

    _auto_width(ws)

    # ── Sheet 6 : Reviews ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Reviews")
    headers = [
        "Review ID", "Order ID", "Customer ID", "Product",
        "Rating (1-5)", "Title", "Review Text", "Review Date",
        "Return Request", "Return Reason", "Return Details", "Return Status",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        _hdr(ws.cell(1, col))
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for rev in sorted(reviews, key=lambda x: x.get("review_date", ""), reverse=True):
        row = [
            rev["id"],
            rev.get("order_id", ""),
            rev.get("customer_id", ""),
            prod_map.get(rev.get("product_id"), "Unknown"),
            rev.get("rating", ""),
            rev.get("title", ""),
            rev.get("review_text", ""),
            rev.get("review_date", ""),
            "Yes" if rev.get("has_return_request") else "No",
            rev.get("return_reason", "") or "",
            rev.get("return_reason_details", "") or "",
            rev.get("return_status", "") or "",
        ]
        ws.append(row)
        r = ws.max_row
        for col in range(1, len(headers) + 1):
            _body(ws.cell(r, col), bold=(col == 1))

    _auto_width(ws)

    # ── Sheet 7 : Sales Summary ───────────────────────────────────────────────
    ws = wb.create_sheet("Sales Summary")
    if sales:
        headers = list(sales[0].keys())
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            _hdr(ws.cell(1, col))
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        for entry in sales:
            row = [entry.get(k, "") for k in headers]
            ws.append(row)
            r = ws.max_row
            for col in range(1, len(headers) + 1):
                _body(ws.cell(r, col), bold=(col == 1))

        _auto_width(ws)

    # ── Stream the workbook ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"AskMyStore_Export_{date.today().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
